from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response, send_file
from datetime import datetime, timedelta
import mysql.connector
from functools import wraps
import hashlib
import os
import csv
from io import StringIO, BytesIO
from werkzeug.utils import secure_filename

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. PDF exports will not work. Run: pip install reportlab")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel exports will not work. Run: pip install openpyxl")


app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

# Database configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', '1234'),  
    'database': os.getenv('DB_NAME', 'HRMS')
}

# File upload configuration
UPLOAD_FOLDER = 'uploads/documents'
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db_connection():
    """Create and return a database connection"""
    return mysql.connector.connect(**DB_CONFIG)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in first', 'warning')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'role' not in session:
                return redirect(url_for('index'))
            if session['role'] not in roles:
                flash('Access denied', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_time_ago(date):
    """Calculate time ago from a date"""
    if not date:
        return 'Unknown'
    
    now = datetime.now().date()
    if isinstance(date, datetime):
        date = date.date()
    
    diff = (now - date).days
    
    if diff == 0:
        return 'Today'
    elif diff == 1:
        return 'Yesterday'
    elif diff < 7:
        return f'{diff} days ago'
    elif diff < 30:
        weeks = diff // 7
        return f'{weeks} week{"s" if weeks > 1 else ""} ago'
    elif diff < 365:
        months = diff // 30
        return f'{months} month{"s" if months > 1 else ""} ago'
    else:
        years = diff // 365
        return f'{years} year{"s" if years > 1 else ""} ago'

# Duplicate role_required definition removed
# def role_required(*roles):
#     def decorator(f):
#         @wraps(f)
#         def decorated_function(*args, **kwargs):
#             if 'role' not in session:
#                 return redirect(url_for('index'))
#             if session['role'] not in roles:
#                 flash('Access denied', 'danger')
#                 return redirect(url_for('dashboard'))
#             return f(*args, **kwargs)
#         return decorated_function
#     return decorator

@app.route('/')
def index():
    if 'user_id' in session:
        if session['role'] == 'HR':
            return redirect(url_for('hr_dashboard'))
        else:
            return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username')
    password = request.form.get('password')
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Hash password using SHA256 to match database
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        cursor.execute(
            "SELECT u.user_id, u.username, u.password_hash, u.role, e.first_name, e.last_name "
            "FROM Users u "
            "LEFT JOIN Employees e ON u.user_id = e.user_id "
            "WHERE u.username = %s",
            (username,)
        )
        user = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if user and user['password_hash'] == password_hash:
            # Store user info in session
            session['user_id'] = user['user_id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['first_name'] = user['first_name'] or 'User'
            session['last_name'] = user['last_name'] or ''
            
            if user['role'] == 'HR':
                return redirect(url_for('hr_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
            return redirect(url_for('index'))
            
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('index'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    """Password reset using email verification only"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        
        if not email:
            flash('Please enter your email address.', 'error')
            return render_template('forgot_password.html')
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            
            cursor.execute("""
                SELECT employee_id, first_name, last_name, email, user_id
                FROM Employees
                WHERE email = %s
            """, (email,))
            
            employee = cursor.fetchone()
            
            if not employee:
                flash('If an account exists with this email, password reset instructions have been prepared.', 'info')
                cursor.close()
                conn.close()
                return render_template('forgot_password.html')
            
            if not employee['user_id']:
                flash('If an account exists with this email, password reset instructions have been prepared.', 'info')
                cursor.close()
                conn.close()
                return render_template('forgot_password.html')
            
            # Get username
            cursor.execute("""
                SELECT username FROM Users WHERE user_id = %s
            """, (employee['user_id'],))
            
            user = cursor.fetchone()
            
            if user:
                # Store user_id in session for password reset
                session['reset_user_id'] = employee['user_id']
                session['reset_username'] = user['username']
                session['reset_expiry'] = (datetime.now() + timedelta(minutes=15)).isoformat()
                
                flash(f'Identity verified! You can now reset your password, {employee["first_name"]}.', 'success')
                cursor.close()
                conn.close()
                return redirect(url_for('reset_password_direct'))
            else:
                flash('If an account exists with this email, password reset instructions have been prepared.', 'info')
                cursor.close()
                conn.close()
        
        except mysql.connector.Error as err:
            flash('An error occurred. Please try again later.', 'error')
            if conn:
                conn.close()
        except Exception as e:
            flash('An unexpected error occurred. Please try again.', 'error')
    
    return render_template('forgot_password.html')

@app.route('/reset_password_direct', methods=['GET', 'POST'])
def reset_password_direct():
    """Reset password directly after identity verification"""
    
    # Check if user has verified their identity
    if 'reset_user_id' not in session:
        flash('Please verify your identity first.', 'warning')
        return redirect(url_for('forgot_password'))
    
    # Check if reset token has expired (15 minutes)
    try:
        expiry = datetime.fromisoformat(session['reset_expiry'])
        if datetime.now() > expiry:
            session.pop('reset_user_id', None)
            session.pop('reset_username', None)
            session.pop('reset_expiry', None)
            flash('Password reset session expired. Please try again.', 'error')
            return redirect(url_for('forgot_password'))
    except:
        flash('Invalid reset session. Please try again.', 'error')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        new_password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        # Validate passwords
        if not new_password or len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('reset_password.html', username=session.get('reset_username'))
        
        if new_password != confirm_password:
            flash('Passwords do not match. Please try again.', 'error')
            return render_template('reset_password.html', username=session.get('reset_username'))
        
        try:
            # Hash the new password
            hashed_pw = hashlib.sha256(new_password.encode()).hexdigest()
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Update password
            cursor.execute("""
                UPDATE Users 
                SET password_hash = %s 
                WHERE user_id = %s
            """, (hashed_pw, session['reset_user_id']))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            # Clear reset session
            session.pop('reset_user_id', None)
            session.pop('reset_username', None)
            session.pop('reset_expiry', None)
            
            flash('Your password has been successfully reset! Please log in with your new password.', 'success')
            return redirect(url_for('index'))
            
        except mysql.connector.Error as err:
            flash(f'Database error: {err}', 'error')
            if conn:
                conn.close()
    
    return render_template('reset_password.html', username=session.get('reset_username'))
@app.route('/dashboard')
@login_required
def dashboard():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Get employee info
        cursor.execute("""
            SELECT e.employee_id, e.first_name, e.last_name, e.position,
                   e.annual_leave_days, e.sick_leave_days, e.leave_days_used
            FROM Employees e
            JOIN Users u ON e.user_id = u.user_id
            WHERE u.user_id = %s
        """, (session['user_id'],))
        employee = cursor.fetchone()

        if not employee:
            flash("Employee record not found.", "error")
            return redirect(url_for('logout'))

        employee_id = employee["employee_id"]

        # Calculate remaining leave days
        total_annual_leave = employee['annual_leave_days']
        used_leaves = float(employee['leave_days_used'])
        remaining_leaves = total_annual_leave - used_leaves

        # Get recent leave requests
        cursor.execute("""
            SELECT leave_type, reason, start_date, end_date, status, days_requested
            FROM LeaveRequests
            WHERE employee_id = %s
            ORDER BY leave_id DESC LIMIT 5
        """, (employee_id,))
        leave_activities = cursor.fetchall()

        # Format leave activities for display
        recent_activities = []
        for leave in leave_activities:
            recent_activities.append({
                'activity_type': f'{leave["leave_type"]} Leave',
                'details': leave['reason'] or 'No reason provided',
                'date': leave['start_date'],
                'status': leave['status'],
                'days': leave['days_requested']
            })

        cursor.close()
        conn.close()

        # Choose dashboard based on position
        position = employee["position"].lower()

        if "driver" in position or "courier" in position:
            template = "dashboard.html"
        else:
            template = "staffDashboard.html"

        return render_template(
            template,
            user=employee,
            remaining_leaves=remaining_leaves,
            recent_activities=recent_activities,
        )

    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('index'))

@app.route('/hr-dashboard')
@login_required
@role_required('HR')
def hr_dashboard():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Total employees count
        cursor.execute("SELECT COUNT(*) as total_employees FROM Employees")
        total_employees = cursor.fetchone()['total_employees']
        
        # Pending leave requests count
        cursor.execute("SELECT COUNT(*) as pending_leaves FROM LeaveRequests WHERE status = 'Pending'")
        pending_leaves = cursor.fetchone()['pending_leaves']
        
        # Expiring documents count (within 60 days)
        cursor.execute("""
            SELECT COUNT(*) as expiring_docs 
            FROM Documents 
            WHERE expiry_date IS NOT NULL 
            AND expiry_date <= DATE_ADD(CURDATE(), INTERVAL 60 DAY)
            AND expiry_date >= CURDATE()
        """)
        expiring_docs = cursor.fetchone()['expiring_docs']
        
        # Active drivers count
        cursor.execute("""
            SELECT COUNT(*) as active_drivers 
            FROM Employees 
            WHERE position LIKE '%Driver%' OR position LIKE '%Courier%'
        """)
        active_drivers = cursor.fetchone()['active_drivers']
        
        # Recent leave requests (last 10)
        cursor.execute("""
            SELECT lr.leave_id, e.first_name, e.last_name, lr.leave_type,
                   lr.start_date, lr.end_date, lr.days_requested,
                   lr.reason, lr.status 
            FROM LeaveRequests lr 
            JOIN Employees e ON lr.employee_id = e.employee_id 
            ORDER BY lr.leave_id DESC LIMIT 10
        """)
        recent_leaves = cursor.fetchall()
        
        # Expiring documents (next 60 days, top 5)
        cursor.execute("""
            SELECT d.doc_type, d.expiry_date, e.first_name, e.last_name,
                   DATEDIFF(d.expiry_date, CURDATE()) as days_until_expiry
            FROM Documents d
            JOIN Employees e ON d.employee_id = e.employee_id
            WHERE d.expiry_date IS NOT NULL
            AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 60 DAY)
            ORDER BY d.expiry_date ASC
            LIMIT 5
        """)
        expiring_documents = cursor.fetchall()
        
        # Employee distribution by position
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN position LIKE '%Driver%' OR position LIKE '%Courier%' THEN 'Courier Drivers'
                    WHEN position LIKE '%Manager%' OR position LIKE '%Supervisor%' THEN 'Management'
                    WHEN position LIKE '%HR%' THEN 'HR Staff'
                    ELSE 'Administrative Staff'
                END as position,
                COUNT(*) as count
            FROM Employees
            GROUP BY 
                CASE 
                    WHEN position LIKE '%Driver%' OR position LIKE '%Courier%' THEN 'Courier Drivers'
                    WHEN position LIKE '%Manager%' OR position LIKE '%Supervisor%' THEN 'Management'
                    WHEN position LIKE '%HR%' THEN 'HR Staff'
                    ELSE 'Administrative Staff'
                END
            ORDER BY count DESC
        """)
        distribution = cursor.fetchall()
        
        # Calculate percentages
        employee_distribution = []
        for item in distribution:
            percentage = (item['count'] / total_employees * 100) if total_employees > 0 else 0
            employee_distribution.append({
                'position': item['position'],
                'count': item['count'],
                'percentage': round(percentage, 1)
            })
        
        # Recent activities (last 10 actions)
        recent_activities = []
        
        # Approved leaves
        cursor.execute("""
            SELECT CONCAT(e.first_name, ' ', e.last_name) as name, 
                   lr.approval_date, 'Leave Approved' as action
            FROM LeaveRequests lr
            JOIN Employees e ON lr.employee_id = e.employee_id
            WHERE lr.status = 'Approved' AND lr.approval_date IS NOT NULL
            ORDER BY lr.approval_date DESC
            LIMIT 5
        """)
        approved_leaves = cursor.fetchall()
        
        for leave in approved_leaves:
            time_ago = get_time_ago(leave['approval_date'])
            recent_activities.append({
                'action': 'Leave Approved',
                'details': leave['name'],
                'time_ago': time_ago,
                'icon': 'fa-check-circle',
                'color': '#4CAF50'
            })
        
        # Recently added employees
        cursor.execute("""
            SELECT CONCAT(first_name, ' ', last_name) as name, hire_date
            FROM Employees
            ORDER BY hire_date DESC
            LIMIT 3
        """)
        new_employees = cursor.fetchall()
        
        for emp in new_employees:
            time_ago = get_time_ago(emp['hire_date'])
            recent_activities.append({
                'action': 'New Employee Added',
                'details': emp['name'],
                'time_ago': time_ago,
                'icon': 'fa-user-plus',
                'color': '#2196F3'
            })
        
        # Sort activities by most recent
        recent_activities.sort(key=lambda x: x['time_ago'])
        recent_activities = recent_activities[:10]
        
        cursor.close()
        conn.close()
        
        # Current time
        current_time = datetime.now().strftime('%d %B %Y, %H:%M')
        
        return render_template(
            'hrDashboard.html',
            user=session,
            total_employees=total_employees,
            pending_leaves=pending_leaves,
            expiring_docs=expiring_docs,
            active_drivers=active_drivers,
            recent_leaves=recent_leaves,
            expiring_documents=expiring_documents,
            employee_distribution=employee_distribution,
            recent_activities=recent_activities,
            current_time=current_time
        )
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('index'))

# get_time_ago function definition is fine.

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('index'))

@app.route('/profile')
@login_required
def profile_page():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee information
        cursor.execute("""
            SELECT e.employee_id, e.first_name, e.last_name, e.email, e.phone, 
                   e.position, e.hire_date, e.annual_leave_days, e.sick_leave_days, 
                   e.leave_days_used
            FROM Employees e
            JOIN Users u ON e.user_id = u.user_id
            WHERE u.user_id = %s
        """, (session['user_id'],))
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('dashboard'))
        
        # Get employee documents
        cursor.execute("""
            SELECT doc_type, issue_date, expiry_date
            FROM Documents
            WHERE employee_id = %s
            ORDER BY issue_date DESC
        """, (employee['employee_id'],))
        raw_documents = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Calculate document status
        today = datetime.now().date()
        thirty_days = today + timedelta(days=30)
        
        documents = []
        for doc in raw_documents:
            status = 'Valid'
            if doc['expiry_date']:
                if doc['expiry_date'] < today:
                    status = 'Expired'
                elif doc['expiry_date'] <= thirty_days:
                    status = 'Expiring'
            
            documents.append({
                'doc_type': doc['doc_type'],
                'issue_date': doc['issue_date'],
                'expiry_date': doc['expiry_date'],
                'status': status
            })
        
        return render_template("profile.html", employee=employee, documents=documents)
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/request')
@login_required
def request_page():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee_id
        cursor.execute(
            "SELECT employee_id, annual_leave_days, sick_leave_days, leave_days_used FROM Employees WHERE user_id = %s",
            (session['user_id'],)
        )
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('dashboard'))
        
        employee_id = employee['employee_id']
        
        # Calculate leave balance
        leave_balance = {
            'annual_leave_days': employee['annual_leave_days'],
            'sick_leave_days': employee['sick_leave_days'],
            'used': float(employee['leave_days_used']),
            'remaining_annual': employee['annual_leave_days'] - float(employee['leave_days_used'])
        }
        
        # Get pending requests count
        cursor.execute(
            "SELECT COUNT(*) as pending_count FROM LeaveRequests WHERE employee_id = %s AND status = 'Pending'",
            (employee_id,)
        )
        pending_count = cursor.fetchone()['pending_count']
        
        # Get leave history
        cursor.execute("""
            SELECT leave_id, leave_type, start_date, end_date, days_requested, 
                   reason, status, approval_date
            FROM LeaveRequests
            WHERE employee_id = %s
            ORDER BY leave_id DESC
        """, (employee_id,))
        leave_history = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Get today's date for date picker validation
        today = datetime.now().strftime('%Y-%m-%d')
        
        return render_template(
            'request.html', 
            user=session, 
            leave_balance=leave_balance,
            pending_count=pending_count,
            leave_history=leave_history,
            today=today
        )
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/submit-leave', methods=['POST'])
@login_required
def submit_leave():
    """Submit a new leave request"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee_id
        cursor.execute(
            "SELECT employee_id FROM Employees WHERE user_id = %s",
            (session['user_id'],)
        )
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('request_page'))
        
        employee_id = employee['employee_id']
        leave_type = request.form.get('leave_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        reason = request.form.get('reason')
        
        # Calculate days requested
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        days_requested = (end - start).days + 1
        
        cursor.execute("""
            INSERT INTO LeaveRequests (employee_id, leave_type, start_date, end_date, 
                                       days_requested, reason, status)
            VALUES (%s, %s, %s, %s, %s, %s, 'Pending')
        """, (employee_id, leave_type, start_date, end_date, days_requested, reason))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Leave request submitted successfully!', 'success')
        return redirect(url_for('dashboard'))
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('request_page'))

@app.route('/notifications')
@login_required
def notifications_page():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee_id
        cursor.execute(
            "SELECT employee_id FROM Employees WHERE user_id = %s",
            (session['user_id'],)
        )
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('dashboard'))
        
        employee_id = employee['employee_id']
        
        # Get leave request notifications (recent approved/rejected)
        cursor.execute("""
            SELECT leave_id, leave_type, start_date, end_date, days_requested, 
                   status, approval_date
            FROM LeaveRequests
            WHERE employee_id = %s 
            AND status IN ('Approved', 'Rejected')
            AND approval_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
            ORDER BY approval_date DESC
        """, (employee_id,))
        leave_notifications = cursor.fetchall()
        
        # Get expiring documents with days calculation
        cursor.execute("""
            SELECT d.doc_type, d.expiry_date,
                   DATEDIFF(d.expiry_date, CURDATE()) as days_until_expiry
            FROM Documents d
            WHERE d.employee_id = %s 
            AND d.expiry_date IS NOT NULL
            AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)
            ORDER BY d.expiry_date ASC
        """, (employee_id,))
        expiring_documents = cursor.fetchall()
        
        # Count statistics
        expiring_docs_count = len([d for d in expiring_documents if d['days_until_expiry'] <= 30])
        leave_updates_count = len(leave_notifications)
        total_notifications = len(leave_notifications) + len(expiring_documents)
        
        cursor.close()
        conn.close()
        
        # System alerts (can be expanded later for actual system announcements)
        system_alerts = []
        
        return render_template(
            'notifications.html', 
            user=session, 
            leave_notifications=leave_notifications,
            expiring_documents=expiring_documents,
            expiring_docs_count=expiring_docs_count,
            leave_updates_count=leave_updates_count,
            total_notifications=total_notifications,
            system_alerts=system_alerts
        )
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/documents')
@login_required
def documents_page():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute(
            "SELECT employee_id FROM Employees WHERE user_id = %s",
            (session['user_id'],)
        )
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('dashboard'))
        
        employee_id = employee['employee_id']
        
        cursor.execute(
            "SELECT document_id, doc_type, file_path, issue_date, expiry_date "
            "FROM Documents "
            "WHERE employee_id = %s "
            "ORDER BY issue_date DESC",
            (employee_id,)
        )
        raw_documents = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        today = datetime.now().date()
        thirty_days = today + timedelta(days=30)
        
        documents = []
        valid_count = 0
        expiring_count = 0
        
        for doc in raw_documents:
            file_name = os.path.basename(doc['file_path']) if doc['file_path'] else 'N/A'
            
            status = 'Valid'
            if doc['expiry_date']:
                if doc['expiry_date'] < today:
                    status = 'Expired'
                elif doc['expiry_date'] <= thirty_days:
                    status = 'Expiring'
                    expiring_count += 1
                else:
                    valid_count += 1
            else:
                valid_count += 1
            
            documents.append({
                'id': doc['document_id'],
                'document_type': doc['doc_type'],
                'file_name': file_name,
                'file_path': doc['file_path'],
                'upload_date': doc['issue_date'],
                'expiry_date': doc['expiry_date'],
                'status': status
            })
        
        total_documents = len(documents)
        
        return render_template(
            'documents.html',
            user=session,
            documents=documents,
            total_documents=total_documents,
            valid_documents=valid_count,
            expiring_soon=expiring_count
        )
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/documents/upload', methods=['POST'])
@login_required
def upload_document():
    flash('Document upload functionality coming soon', 'info')
    return redirect(url_for('documents_page'))

@app.route('/documents/view/<int:doc_id>')
@login_required
def view_document(doc_id):
    flash('Document viewing functionality coming soon', 'info')
    return redirect(url_for('documents_page'))

@app.route('/documents/download/<int:doc_id>')
@login_required
def download_document(doc_id):
    flash('Document download functionality coming soon', 'info')
    return redirect(url_for('documents_page'))



# ================================
# HR MANAGEMENT PAGES
# ================================

@app.route('/manageEmployees')
@login_required
@role_required('HR')
def manageEmployees():
    """Display all employees for HR management"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT e.employee_id, e.first_name, e.last_name, e.email, e.phone, 
                   e.position, e.hire_date, e.annual_leave_days, e.sick_leave_days, 
                   e.leave_days_used
            FROM Employees e
            ORDER BY e.employee_id
        """)
        employees = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Get today's date for the form
        today = datetime.now().strftime('%Y-%m-%d')
        
        return render_template('manageEmployees.html', employees=employees, today=today)
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('hr_dashboard'))

@app.route('/addEmployees')
@login_required
@role_required('HR')
def addEmployees():
    """Display the add employee form"""
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('addEmployees.html', today=today)

@app.route('/addEmployee', methods=['POST'])
@login_required
@role_required('HR')
def addEmployee():
    """Add a new employee record with login credentials"""
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    position = request.form.get('position')
    hire_date = request.form.get('hire_date')
    annual_leave_days = request.form.get('annual_leave_days', 15)
    sick_leave_days = request.form.get('sick_leave_days', 10)
    
    # Login credentials
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role', 'Employee')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if username already exists
        cursor.execute("SELECT user_id FROM Users WHERE username = %s", (username,))
        if cursor.fetchone():
            flash(f'Username "{username}" already exists. Please choose a different username.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('addEmployees'))
        
        # Check if email already exists
        cursor.execute("SELECT employee_id FROM Employees WHERE email = %s", (email,))
        if cursor.fetchone():
            flash(f'Email "{email}" already exists in the system.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('addEmployees'))
        
        # Hash the password using SHA256
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Insert into Users table first
        cursor.execute("""
            INSERT INTO Users (username, password_hash, role)
            VALUES (%s, %s, %s)
        """, (username, password_hash, role))
        
        # Get the newly created user_id
        user_id = cursor.lastrowid
        
        # Insert into Employees table with the user_id
        cursor.execute("""
            INSERT INTO Employees (first_name, last_name, email, phone, position, hire_date,
                                   annual_leave_days, sick_leave_days, leave_days_used, user_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, %s)
        """, (first_name, last_name, email, phone, position, hire_date, 
              annual_leave_days, sick_leave_days, user_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f'Employee {first_name} {last_name} added successfully! Username: {username}', 'success')
    except mysql.connector.Error as err:
        flash(f"Error adding employee: {err}", "error")
        if conn:
            conn.close()

    return redirect(url_for('manageEmployees'))

@app.route('/deleteEmployee/<int:employee_id>', methods=['POST'])
@login_required
@role_required('HR')
def deleteEmployee(employee_id):
    """Delete an employee record"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM Employees WHERE employee_id = %s", (employee_id,))
        conn.commit()
        flash('Employee deleted successfully!', 'info')
    except mysql.connector.Error as err:
        flash(f"Error deleting employee: {err}", "error")
    finally:
        conn.close()

    return redirect(url_for('manageEmployees'))

# Add these routes to your app.py file

@app.route('/editEmployee/<int:employee_id>')
@login_required
@role_required('HR')
def editEmployee(employee_id):
    """Display the edit employee form"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee details
        cursor.execute("""
            SELECT e.employee_id, e.first_name, e.last_name, e.email, e.phone, 
                   e.position, e.hire_date, e.annual_leave_days, e.sick_leave_days, 
                   e.leave_days_used, e.user_id, u.username, u.role
            FROM Employees e
            LEFT JOIN Users u ON e.user_id = u.user_id
            WHERE e.employee_id = %s
        """, (employee_id,))
        
        employee = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not employee:
            flash('Employee not found', 'error')
            return redirect(url_for('manageEmployees'))
        
        # Get today's date for form validation
        today = datetime.now().strftime('%Y-%m-%d')
        
        return render_template('editEmployee.html', employee=employee, today=today)
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('manageEmployees'))


@app.route('/updateEmployee/<int:employee_id>', methods=['POST'])
@login_required
@role_required('HR')
def updateEmployee(employee_id):
    """Update an employee record"""
    try:
        # Get form data
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        position = request.form.get('position')
        hire_date = request.form.get('hire_date')
        annual_leave_days = request.form.get('annual_leave_days')
        sick_leave_days = request.form.get('sick_leave_days')
        role = request.form.get('role')
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if email already exists for another employee
        cursor.execute("""
            SELECT employee_id FROM Employees 
            WHERE email = %s AND employee_id != %s
        """, (email, employee_id))
        
        if cursor.fetchone():
            flash(f'Email "{email}" is already used by another employee.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('editEmployee', employee_id=employee_id))
        
        # Update employee information
        cursor.execute("""
            UPDATE Employees 
            SET first_name = %s, last_name = %s, email = %s, phone = %s, 
                position = %s, hire_date = %s, annual_leave_days = %s, sick_leave_days = %s
            WHERE employee_id = %s
        """, (first_name, last_name, email, phone, position, hire_date, 
              annual_leave_days, sick_leave_days, employee_id))
        
        # Update user role if user_id exists
        cursor.execute("SELECT user_id FROM Employees WHERE employee_id = %s", (employee_id,))
        result = cursor.fetchone()
        
        if result and result['user_id']:
            cursor.execute("""
                UPDATE Users 
                SET role = %s 
                WHERE user_id = %s
            """, (role, result['user_id']))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f'Employee {first_name} {last_name} updated successfully!', 'success')
        return redirect(url_for('manageEmployees'))
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('editEmployee', employee_id=employee_id))

@app.route('/leave-management')
@login_required
@role_required('HR')
def leaveManagement():
    """HR - View and manage leave requests"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Get all leave requests with employee details
        cursor.execute("""
            SELECT lr.leave_id, lr.employee_id, e.first_name, e.last_name,
                   lr.leave_type, lr.start_date, lr.end_date, lr.days_requested,
                   lr.reason, lr.status, lr.approval_date
            FROM LeaveRequests lr
            JOIN Employees e ON lr.employee_id = e.employee_id
            ORDER BY 
                CASE lr.status 
                    WHEN 'Pending' THEN 1 
                    WHEN 'Approved' THEN 2 
                    WHEN 'Rejected' THEN 3 
                END,
                lr.start_date DESC
        """)
        leaves = cursor.fetchall()

        # Get summary statistics
        cursor.execute("SELECT COUNT(*) as total FROM LeaveRequests")
        total_count = cursor.fetchone()['total']

        cursor.execute("SELECT COUNT(*) as pending FROM LeaveRequests WHERE status = 'Pending'")
        pending_count = cursor.fetchone()['pending']

        cursor.execute("""
            SELECT COUNT(*) as approved 
            FROM LeaveRequests 
            WHERE status = 'Approved' 
            AND approval_date >= DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
        """)
        approved_count = cursor.fetchone()['approved']

        cursor.execute("""
            SELECT COUNT(*) as rejected 
            FROM LeaveRequests 
            WHERE status = 'Rejected' 
            AND approval_date >= DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
        """)
        rejected_count = cursor.fetchone()['rejected']

        cursor.close()
        conn.close()

        return render_template(
            'leaveManagement.html', 
            user=session, 
            leaves=leaves,
            total_count=total_count,
            pending_count=pending_count,
            approved_count=approved_count,
            rejected_count=rejected_count
        )

    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('hr_dashboard'))

@app.route('/approve-leave/<int:leave_id>', methods=['POST'])
@login_required
@role_required('HR')
def approve_leave(leave_id):
    """Approve a leave request"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get leave request details
        cursor.execute("""
            SELECT employee_id, days_requested 
            FROM LeaveRequests 
            WHERE leave_id = %s
        """, (leave_id,))
        leave_req = cursor.fetchone()
        
        # Update leave request status
        cursor.execute("""
            UPDATE LeaveRequests 
            SET status = 'Approved', 
                approved_by = %s,
                approval_date = CURDATE()
            WHERE leave_id = %s
        """, (session['user_id'], leave_id))
        
        # Update employee's used leave days
        cursor.execute("""
            UPDATE Employees 
            SET leave_days_used = leave_days_used + %s
            WHERE employee_id = %s
        """, (leave_req['days_requested'], leave_req['employee_id']))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Leave request approved successfully!', 'success')
    except mysql.connector.Error as err:
        flash(f"Error approving leave: {err}", "error")
    
    return redirect(url_for('leaveManagement'))

@app.route('/reject-leave/<int:leave_id>', methods=['POST'])
@login_required
@role_required('HR')
def reject_leave(leave_id):
    """Reject a leave request"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE LeaveRequests 
            SET status = 'Rejected', 
                approved_by = %s,
                approval_date = CURDATE()
            WHERE leave_id = %s
        """, (session['user_id'], leave_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Leave request rejected.', 'info')
    except mysql.connector.Error as err:
        flash(f"Error rejecting leave: {err}", "error")
    
    return redirect(url_for('leaveManagement'))

@app.route('/document-management')
@login_required
@role_required('HR')
def documentManagement():
    """HR - Manage employee documents"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Get all documents with employee details and expiry calculations
        cursor.execute("""
            SELECT d.document_id, d.employee_id, e.first_name, e.last_name, e.position,
                   d.doc_type, d.issue_date, d.expiry_date, d.file_path,
                   DATEDIFF(d.expiry_date, CURDATE()) as days_until_expiry
            FROM Documents d
            JOIN Employees e ON d.employee_id = e.employee_id
            ORDER BY 
                CASE 
                    WHEN d.expiry_date IS NULL THEN 3
                    WHEN d.expiry_date < CURDATE() THEN 1
                    WHEN d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 2
                    ELSE 4
                END,
                d.expiry_date ASC
        """)
        raw_documents = cursor.fetchall()

        cursor.close()
        conn.close()

        # Process documents and add status classification
        documents = []
        total_documents = len(raw_documents)
        valid_documents = 0
        expiring_soon = 0
        expired_documents = 0

        for doc in raw_documents:
            # Determine status class for filtering
            if doc['days_until_expiry'] is None:
                status_class = 'valid'
                valid_documents += 1
            elif doc['days_until_expiry'] < 0:
                status_class = 'expired'
                expired_documents += 1
            elif doc['days_until_expiry'] <= 30:
                status_class = 'expiring'
                expiring_soon += 1
            else:
                status_class = 'valid'
                valid_documents += 1

            documents.append({
                'document_id': doc['document_id'],
                'employee_id': doc['employee_id'],
                'first_name': doc['first_name'],
                'last_name': doc['last_name'],
                'position': doc['position'],
                'doc_type': doc['doc_type'],
                'issue_date': doc['issue_date'],
                'expiry_date': doc['expiry_date'],
                'file_path': doc['file_path'],
                'days_until_expiry': doc['days_until_expiry'],
                'status_class': status_class
            })

        return render_template(
            'documentManagement.html', 
            user=session, 
            documents=documents,
            total_documents=total_documents,
            valid_documents=valid_documents,
            expiring_soon=expiring_soon,
            expired_documents=expired_documents
        )

    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('hr_dashboard'))


# Add these routes to your app.py file

# File upload configuration
UPLOAD_FOLDER = 'uploads/documents'
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# HR - Upload document for any employee
@app.route('/upload-document-hr')
@login_required
@role_required('HR')
def uploadDocumentHR():
    """Display upload form for HR staff"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get all employees for dropdown
        cursor.execute("""
            SELECT employee_id, first_name, last_name, position
            FROM Employees
            ORDER BY first_name, last_name
        """)
        employees = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template('uploadDocumentHR.html', employees=employees)
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('documentManagement'))

@app.route('/submit-document-hr', methods=['POST'])
@login_required
@role_required('HR')
def submit_document_hr():
    """Process document upload by HR staff"""
    try:
        employee_id = request.form.get('employee_id')
        doc_type = request.form.get('doc_type')
        issue_date = request.form.get('issue_date')
        expiry_date = request.form.get('expiry_date')
        
        # Handle file upload
        if 'document_file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('uploadDocument'))
        
        file = request.files['document_file']
        
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('uploadDocumentHR'))
        
        if not allowed_file(file.filename):
            flash('Invalid file type. Allowed types: PDF, JPG, JPEG, PNG, DOC, DOCX', 'error')
            return redirect(url_for('uploadDocumentHR'))
        
        # Secure the filename and save
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{employee_id}_{doc_type}_{timestamp}_{filename}"
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        
        file.save(file_path)
        
        # Insert into database
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Handle expiry date - set to NULL if not applicable
        if not expiry_date or doc_type == 'ID':
            expiry_date = None
        
        cursor.execute("""
            INSERT INTO Documents (employee_id, doc_type, file_path, issue_date, expiry_date)
            VALUES (%s, %s, %s, %s, %s)
        """, (employee_id, doc_type, file_path, issue_date, expiry_date))
        
        # Create expiry alert if applicable
        if expiry_date:
            document_id = cursor.lastrowid
            cursor.execute("""
                INSERT INTO ExpiryAlerts (document_id, alert_date, status)
                VALUES (%s, DATE_SUB(%s, INTERVAL 60 DAY), 'Unread')
            """, (document_id, expiry_date))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Document uploaded successfully!', 'success')
        return redirect(url_for('documentManagement'))
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('uploadDocumentHR'))
    except Exception as e:
        flash(f'Error uploading file: {str(e)}', 'error')
        return redirect(url_for('uploadDocumentHR'))

# Employee - Upload their own document
@app.route('/upload-document-employee')
@login_required
def uploadDocument():
    """Display upload form for employees"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee info
        cursor.execute("""
            SELECT employee_id, first_name, last_name
            FROM Employees
            WHERE user_id = %s
        """, (session['user_id'],))
        employee = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('documents_page'))
        
        return render_template('uploadDocument.html', employee=employee)
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('documents_page'))

@app.route('/submit-document-employee', methods=['POST'])
@login_required
def submit_document_employee():
    """Process document upload by employee"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get employee_id
        cursor.execute(
            "SELECT employee_id FROM Employees WHERE user_id = %s",
            (session['user_id'],)
        )
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee record not found', 'error')
            return redirect(url_for('documents_page'))
        
        employee_id = employee['employee_id']
        doc_type = request.form.get('doc_type')
        issue_date = request.form.get('issue_date')
        expiry_date = request.form.get('expiry_date')
        
        # Handle file upload
        if 'document_file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('upload_document_employee'))
        
        file = request.files['document_file']
        
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('upload_document_employee'))
        
        if not allowed_file(file.filename):
            flash('Invalid file type. Allowed types: PDF, JPG, JPEG, PNG, DOC, DOCX', 'error')
            return redirect(url_for('upload_document_employee'))
        
        # Secure the filename and save
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{employee_id}_{doc_type}_{timestamp}_{filename}"
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        
        file.save(file_path)
        
        # Handle expiry date
        if not expiry_date or doc_type == 'ID':
            expiry_date = None
        
        cursor.execute("""
            INSERT INTO Documents (employee_id, doc_type, file_path, issue_date, expiry_date)
            VALUES (%s, %s, %s, %s, %s)
        """, (employee_id, doc_type, file_path, issue_date, expiry_date))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Document uploaded successfully!', 'success')
        return redirect(url_for('documents_page'))
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('upload_document_employee'))
    except Exception as e:
        flash(f'Error uploading file: {str(e)}', 'error')
        return redirect(url_for('upload_document_employee'))
    
@app.route('/expiry-alerts')
@login_required
@role_required('HR')
def expiryAlerts():
    """HR - View all expiry alerts with priority classification and enhanced filtering"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Get filter parameter from query string
        filter_type = request.args.get('filter', 'all')
        
        # Base query with all necessary joins
        base_query = """
            SELECT d.document_id, d.employee_id, e.first_name, e.last_name, e.position,
                   d.doc_type, d.expiry_date, d.file_path,
                   DATEDIFF(d.expiry_date, CURDATE()) as days_until_expiry,
                   ea.alert_id, ea.alert_date, ea.status
            FROM Documents d
            JOIN Employees e ON d.employee_id = e.employee_id
            LEFT JOIN ExpiryAlerts ea ON d.document_id = ea.document_id
            WHERE d.expiry_date IS NOT NULL
        """
        
        # Apply filters based on filter_type
        if filter_type == 'critical':
            base_query += " AND d.expiry_date < CURDATE()"
        elif filter_type == 'urgent':
            base_query += " AND d.expiry_date >= CURDATE() AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)"
        elif filter_type == 'upcoming':
            base_query += " AND d.expiry_date > DATE_ADD(CURDATE(), INTERVAL 30 DAY) AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)"
        else:  # 'all'
            base_query += " AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)"
        
        base_query += """
            ORDER BY 
                CASE 
                    WHEN d.expiry_date < CURDATE() THEN 1
                    WHEN d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 2
                    ELSE 3
                END,
                d.expiry_date ASC
        """
        
        cursor.execute(base_query)
        raw_alerts = cursor.fetchall()

        # Process alerts and classify by priority
        alerts = []
        critical_alerts = 0  # Expired
        urgent_alerts = 0    # ≤30 days
        upcoming_alerts = 0  # 31-90 days
        acknowledged_alerts = 0  # Read status

        for alert in raw_alerts:
            days = alert['days_until_expiry']
            
            # Determine priority classification
            if days < 0:
                priority_class = 'critical'
                critical_alerts += 1
            elif days <= 30:
                priority_class = 'urgent'
                urgent_alerts += 1
            else:
                priority_class = 'upcoming'
                upcoming_alerts += 1
            
            # Count acknowledged alerts
            if alert['status'] == 'Read':
                acknowledged_alerts += 1
            
            alerts.append({
                'alert_id': alert['alert_id'],
                'document_id': alert['document_id'],
                'employee_id': alert['employee_id'],
                'first_name': alert['first_name'],
                'last_name': alert['last_name'],
                'position': alert['position'],
                'doc_type': alert['doc_type'],
                'expiry_date': alert['expiry_date'],
                'days_until_expiry': days,
                'alert_date': alert['alert_date'],
                'status': alert['status'] if alert['status'] else 'Unread',
                'priority_class': priority_class,
                'file_path': alert['file_path']
            })

        total_alerts = len(alerts)

        # Get count of documents without alert records (for notification)
        cursor.execute("""
            SELECT COUNT(*) as missing_alerts
            FROM Documents d
            LEFT JOIN ExpiryAlerts ea ON d.document_id = ea.document_id
            WHERE d.expiry_date IS NOT NULL
            AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)
            AND ea.alert_id IS NULL
        """)
        missing_alerts_count = cursor.fetchone()['missing_alerts']

        cursor.close()
        conn.close()

        return render_template(
            'expiryAlerts.html', 
            user=session, 
            alerts=alerts,
            total_alerts=total_alerts,
            critical_alerts=critical_alerts,
            urgent_alerts=urgent_alerts,
            upcoming_alerts=upcoming_alerts,
            acknowledged_alerts=acknowledged_alerts,
            missing_alerts_count=missing_alerts_count,
            current_filter=filter_type
        )

    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('hr_dashboard'))
    except Exception as e:
        flash(f"Unexpected error: {str(e)}", "error")
        return redirect(url_for('hr_dashboard'))


@app.route('/mark-alert-read/<int:alert_id>', methods=['POST'])
@login_required
@role_required('HR')
def mark_alert_read(alert_id):
    """Mark an expiry alert as read/acknowledged"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if alert exists
        cursor.execute("SELECT alert_id, status FROM ExpiryAlerts WHERE alert_id = %s", (alert_id,))
        alert = cursor.fetchone()
        
        if not alert:
            flash('Alert not found.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('expiryAlerts'))
        
        if alert['status'] == 'Read':
            flash('Alert already acknowledged.', 'info')
            cursor.close()
            conn.close()
            return redirect(url_for('expiryAlerts'))
        
        # Update alert status
        cursor.execute("""
            UPDATE ExpiryAlerts 
            SET status = 'Read' 
            WHERE alert_id = %s
        """, (alert_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash('Alert marked as acknowledged successfully.', 'success')
        
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    except Exception as e:
        flash(f"Error updating alert: {str(e)}", "error")
    
    return redirect(url_for('expiryAlerts'))


@app.route('/create-missing-alerts', methods=['POST'])
@login_required
@role_required('HR')
def create_missing_alerts():
    """Create alert records for documents that don't have them"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Find documents with expiry dates but no alerts (within 90 days)
        cursor.execute("""
            INSERT INTO ExpiryAlerts (document_id, alert_date, status)
            SELECT d.document_id, 
                   CASE 
                       WHEN d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 60 DAY) 
                       THEN DATE_SUB(d.expiry_date, INTERVAL 60 DAY)
                       ELSE CURDATE()
                   END as alert_date,
                   'Unread'
            FROM Documents d
            LEFT JOIN ExpiryAlerts ea ON d.document_id = ea.document_id
            WHERE d.expiry_date IS NOT NULL
            AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)
            AND ea.alert_id IS NULL
        """)
        
        rows_added = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
        
        if rows_added > 0:
            flash(f'Successfully created {rows_added} new alert record(s).', 'success')
        else:
            flash('All alerts are already up to date.', 'info')
            
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    except Exception as e:
        flash(f"Error creating alerts: {str(e)}", "error")
    
    return redirect(url_for('expiryAlerts'))


@app.route('/bulk-acknowledge-alerts', methods=['POST'])
@login_required
@role_required('HR')
def bulk_acknowledge_alerts():
    """Acknowledge all unread alerts of a specific priority"""
    try:
        priority = request.form.get('priority', 'all')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if priority == 'critical':
            # Acknowledge all expired document alerts
            cursor.execute("""
                UPDATE ExpiryAlerts ea
                JOIN Documents d ON ea.document_id = d.document_id
                SET ea.status = 'Read'
                WHERE ea.status = 'Unread'
                AND d.expiry_date < CURDATE()
            """)
        elif priority == 'urgent':
            # Acknowledge all urgent (≤30 days) alerts
            cursor.execute("""
                UPDATE ExpiryAlerts ea
                JOIN Documents d ON ea.document_id = d.document_id
                SET ea.status = 'Read'
                WHERE ea.status = 'Unread'
                AND d.expiry_date >= CURDATE()
                AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)
            """)
        elif priority == 'upcoming':
            # Acknowledge all upcoming (31-90 days) alerts
            cursor.execute("""
                UPDATE ExpiryAlerts ea
                JOIN Documents d ON ea.document_id = d.document_id
                SET ea.status = 'Read'
                WHERE ea.status = 'Unread'
                AND d.expiry_date > DATE_ADD(CURDATE(), INTERVAL 30 DAY)
                AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)
            """)
        else:
            # Acknowledge all unread alerts
            cursor.execute("""
                UPDATE ExpiryAlerts
                SET status = 'Read'
                WHERE status = 'Unread'
            """)
        
        rows_updated = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
        
        if rows_updated > 0:
            flash(f'Successfully acknowledged {rows_updated} alert(s).', 'success')
        else:
            flash('No unread alerts to acknowledge.', 'info')
            
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    except Exception as e:
        flash(f"Error acknowledging alerts: {str(e)}", "error")
    
    return redirect(url_for('expiryAlerts'))


@app.route('/view-document/<int:document_id>')
@login_required
@role_required('HR')
def view_document_details(document_id):
    """View detailed information about a specific document"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT d.*, e.first_name, e.last_name, e.position, e.email, e.phone
            FROM Documents d
            JOIN Employees e ON d.employee_id = e.employee_id
            WHERE d.document_id = %s
        """, (document_id,))
        
        document = cursor.fetchone()
        
        if not document:
            flash('Document not found.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('expiryAlerts'))
        
        # Calculate days until expiry
        if document['expiry_date']:
            days_until_expiry = (document['expiry_date'] - datetime.now().date()).days
            document['days_until_expiry'] = days_until_expiry
        
        # Get related alerts
        cursor.execute("""
            SELECT alert_id, alert_date, status
            FROM ExpiryAlerts
            WHERE document_id = %s
            ORDER BY alert_date DESC
        """, (document_id,))
        
        alerts = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template(
            'documentDetails.html',
            document=document,
            alerts=alerts,
            user=session
        )
        
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('expiryAlerts'))
    except Exception as e:
        flash(f"Error retrieving document: {str(e)}", "error")
        return redirect(url_for('expiryAlerts'))


@app.route('/export-alerts/<format>')
@login_required
@role_required('HR')
def export_alerts(format):
    """Export expiry alerts to CSV, Excel, or PDF"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT e.employee_id, e.first_name, e.last_name, e.position,
                   d.doc_type, d.expiry_date,
                   DATEDIFF(d.expiry_date, CURDATE()) as days_until_expiry,
                   ea.alert_date, ea.status
            FROM Documents d
            JOIN Employees e ON d.employee_id = e.employee_id
            LEFT JOIN ExpiryAlerts ea ON d.document_id = ea.document_id
            WHERE d.expiry_date IS NOT NULL
            AND d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 90 DAY)
            ORDER BY d.expiry_date ASC
        """)
        
        alerts = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if format == 'csv':
            return export_alerts_csv(alerts)
        elif format == 'excel':
            if not OPENPYXL_AVAILABLE:
                flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'error')
                return redirect(url_for('expiryAlerts'))
            return export_alerts_excel(alerts)
        elif format == 'pdf':
            if not REPORTLAB_AVAILABLE:
                flash('PDF export requires reportlab. Install with: pip install reportlab', 'error')
                return redirect(url_for('expiryAlerts'))
            return export_alerts_pdf(alerts)
        else:
            flash('Invalid export format', 'error')
            return redirect(url_for('expiryAlerts'))
            
    except Exception as e:
        flash(f'Error exporting alerts: {str(e)}', 'error')
        return redirect(url_for('expiryAlerts'))


def export_alerts_csv(alerts):
    """Export alerts to CSV"""
    output = StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Employee ID', 'Name', 'Position', 'Document Type', 
                     'Expiry Date', 'Days Until Expiry', 'Status', 'Priority'])
    
    for alert in alerts:
        days = alert['days_until_expiry']
        if days < 0:
            priority = 'Critical (Expired)'
        elif days <= 30:
            priority = 'Urgent'
        else:
            priority = 'Upcoming'
        
        writer.writerow([
            alert['employee_id'],
            f"{alert['first_name']} {alert['last_name']}",
            alert['position'],
            alert['doc_type'],
            alert['expiry_date'].strftime('%Y-%m-%d'),
            days,
            alert['status'] or 'Unread',
            priority
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=expiry_alerts_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    response.headers["Content-type"] = "text/csv"
    
    return response


def export_alerts_excel(alerts):
    """Export alerts to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expiry Alerts"
    
    # Styles
    header_fill = PatternFill(start_color='FF4da6d1', end_color='FF4da6d1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    
    # Headers
    headers = ['Employee ID', 'Name', 'Position', 'Document Type', 
               'Expiry Date', 'Days Until Expiry', 'Status', 'Priority']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for alert in alerts:
        days = alert['days_until_expiry']
        if days < 0:
            priority = 'Critical (Expired)'
        elif days <= 30:
            priority = 'Urgent'
        else:
            priority = 'Upcoming'
        
        ws.append([
            alert['employee_id'],
            f"{alert['first_name']} {alert['last_name']}",
            alert['position'],
            alert['doc_type'],
            alert['expiry_date'].strftime('%Y-%m-%d'),
            days,
            alert['status'] or 'Unread',
            priority
        ])
    
    # Auto-adjust columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'expiry_alerts_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def export_alerts_pdf(alerts):
    """Export alerts to PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2d7fa6'),
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("SwiftSend Couriers - Expiry Alerts Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Table
    data = [['Employee', 'Position', 'Doc Type', 'Expiry Date', 'Days', 'Priority']]
    
    for alert in alerts:
        days = alert['days_until_expiry']
        if days < 0:
            priority = 'Critical'
        elif days <= 30:
            priority = 'Urgent'
        else:
            priority = 'Upcoming'
        
        data.append([
            f"{alert['first_name']} {alert['last_name']}"[:20],
            alert['position'][:15],
            alert['doc_type'],
            alert['expiry_date'].strftime('%Y-%m-%d'),
            str(days),
            priority
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4da6d1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'expiry_alerts_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
        mimetype='application/pdf'
    )
# ADD THIS ROUTE TO YOUR app.py (it's missing from your code)

@app.route('/reports')
@login_required
@role_required('HR')
def reportPage():
    """HR - Summary Reports and Analytics"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Total employees count
        cursor.execute("SELECT COUNT(*) AS total_employees FROM Employees")
        total_employees = cursor.fetchone()['total_employees']

        # Pending leave requests
        cursor.execute("SELECT COUNT(*) AS pending_leaves FROM LeaveRequests WHERE status='Pending'")
        pending_leaves = cursor.fetchone()['pending_leaves']

        # Unread expiry alerts
        cursor.execute("SELECT COUNT(*) AS expiring_docs FROM ExpiryAlerts WHERE status='Unread'")
        expiring_docs = cursor.fetchone()['expiring_docs']

        # Recent HR activities (last 15 records)
        cursor.execute("""
            SELECT CONCAT(e.first_name, ' ', e.last_name) AS employee, 
                   lr.leave_type, lr.status, lr.approval_date AS date
            FROM LeaveRequests lr
            JOIN Employees e ON lr.employee_id = e.employee_id
            WHERE lr.status IN ('Approved', 'Rejected')
            ORDER BY lr.leave_id DESC LIMIT 15
        """)
        activities = cursor.fetchall()

        # Leave statistics for current month
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved_count,
                SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected_count
            FROM LeaveRequests 
            WHERE approval_date >= DATE_FORMAT(CURDATE(), '%Y-%m-01')
            AND status IN ('Approved', 'Rejected')
        """)
        leave_stats = cursor.fetchone()
        approved_this_month = leave_stats['approved_count'] or 0
        rejected_this_month = leave_stats['rejected_count'] or 0

        # Total documents tracked
        cursor.execute("SELECT COUNT(*) as total_docs FROM Documents")
        total_documents = cursor.fetchone()['total_docs']

        # Compliance rate calculation
        cursor.execute("""
            SELECT COUNT(DISTINCT d.employee_id) as non_compliant_employees
            FROM Documents d
            WHERE d.expiry_date IS NOT NULL 
            AND (d.expiry_date < CURDATE() 
                 OR d.expiry_date <= DATE_ADD(CURDATE(), INTERVAL 30 DAY))
        """)
        non_compliant = cursor.fetchone()['non_compliant_employees'] or 0
        compliant_employees = total_employees - non_compliant
        compliance_rate = round((compliant_employees / total_employees * 100), 1) if total_employees > 0 else 0

        # Department breakdown
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN position LIKE '%Driver%' OR position LIKE '%Courier%' THEN 'Drivers'
                    WHEN position LIKE '%Manager%' OR position LIKE '%Supervisor%' THEN 'Management'
                    WHEN position LIKE '%HR%' THEN 'Human Resources'
                    WHEN position LIKE '%Customer%' THEN 'Customer Service'
                    WHEN position LIKE '%Admin%' OR position LIKE '%Assistant%' THEN 'Administration'
                    WHEN position LIKE '%Warehouse%' THEN 'Warehouse'
                    ELSE 'Other'
                END as department,
                COUNT(*) as count
            FROM Employees
            GROUP BY department
            ORDER BY count DESC
        """)
        departments = cursor.fetchall()

        # Calculate department percentages
        for dept in departments:
            dept['percentage'] = round((dept['count'] / total_employees * 100), 1) if total_employees > 0 else 0

        cursor.close()
        conn.close()

        return render_template(
            'reportPage.html',
            user=session,
            total_employees=total_employees,
            pending_leaves=pending_leaves,
            expiring_docs=expiring_docs,
            activities=activities,
            approved_this_month=approved_this_month,
            rejected_this_month=rejected_this_month,
            total_documents=total_documents,
            compliance_rate=compliance_rate,
            compliant_employees=compliant_employees,
            non_compliant=non_compliant,
            departments=departments
        )

    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('hr_dashboard'))

# Report Generation Routes - Add these after reportPage()
@app.route('/generate-report/<report_type>')
@login_required
@role_required('HR')
def generate_report(report_type):
    """Generate different types of reports"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        if report_type == 'employee_roster':
            # Get all employee details
            cursor.execute("""
                SELECT e.employee_id, e.first_name, e.last_name, e.position,
                       e.email, e.phone, e.hire_date, e.annual_leave_days, e.sick_leave_days,
                       e.leave_days_used, u.username, u.role
                FROM Employees e
                LEFT JOIN Users u ON e.user_id = u.user_id
                ORDER BY e.last_name, e.first_name
            """)
            data = cursor.fetchall()
            title = "Employee Roster Report"
            template = 'reports/employee_roster.html'
            
        elif report_type == 'leave_summary':
            # Get leave summary for all employees
            cursor.execute("""
                SELECT e.employee_id, e.first_name, e.last_name, e.position,
                       e.annual_leave_days, e.sick_leave_days, e.leave_days_used,
                       (e.annual_leave_days - e.leave_days_used) as remaining_leave,
                       COUNT(lr.leave_id) as total_requests,
                       SUM(CASE WHEN lr.status = 'Pending' THEN 1 ELSE 0 END) as pending_requests,
                       SUM(CASE WHEN lr.status = 'Approved' THEN 1 ELSE 0 END) as approved_requests
                FROM Employees e
                LEFT JOIN LeaveRequests lr ON e.employee_id = lr.employee_id
                GROUP BY e.employee_id
                ORDER BY e.last_name, e.first_name
            """)
            data = cursor.fetchall()
            title = "Leave Summary Report"
            template = 'reports/leave_summary.html'
            
        elif report_type == 'document_compliance':
            # Get document compliance for all employees
            cursor.execute("""
                SELECT e.employee_id, e.first_name, e.last_name, e.position,
                       COUNT(DISTINCT d.document_id) as total_docs,
                       SUM(CASE WHEN d.expiry_date < CURDATE() THEN 1 ELSE 0 END) as expired_docs,
                       SUM(CASE WHEN d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) 
                           THEN 1 ELSE 0 END) as expiring_docs,
                       CASE 
                           WHEN SUM(CASE WHEN d.expiry_date < CURDATE() THEN 1 ELSE 0 END) > 0 THEN 'Non-Compliant'
                           WHEN SUM(CASE WHEN d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) 
                                THEN 1 ELSE 0 END) > 0 THEN 'At Risk'
                           ELSE 'Compliant'
                       END as compliance_status
                FROM Employees e
                LEFT JOIN Documents d ON e.employee_id = d.employee_id
                GROUP BY e.employee_id
                ORDER BY compliance_status, e.last_name, e.first_name
            """)
            data = cursor.fetchall()
            title = "Document Compliance Report"
            template = 'reports/document_compliance.html'
            
        elif report_type == 'monthly_summary':
            # Get monthly summary
            current_month = datetime.now().strftime('%Y-%m')
            
            cursor.execute("SELECT COUNT(*) as total_employees FROM Employees")
            total_employees = cursor.fetchone()['total_employees']
            
            cursor.execute("""
                SELECT COUNT(*) as total_leaves,
                       SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved,
                       SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected,
                       SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending
                FROM LeaveRequests
                WHERE DATE_FORMAT(start_date, '%Y-%m') = %s
            """, (current_month,))
            leave_stats = cursor.fetchone()
            
            cursor.execute("""
                SELECT COUNT(*) as expiring_docs
                FROM Documents
                WHERE expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY)
            """)
            doc_stats = cursor.fetchone()
            
            data = {
                'total_employees': total_employees,
                'leave_stats': leave_stats,
                'doc_stats': doc_stats,
                'month': datetime.now().strftime('%B %Y')
            }
            title = f"Monthly Summary Report - {datetime.now().strftime('%B %Y')}"
            template = 'reports/monthly_summary.html'
        
        else:
            flash('Invalid report type', 'error')
            return redirect(url_for('reportPage'))
        
        cursor.close()
        conn.close()
        
        return render_template(template, data=data, title=title, generated_date=datetime.now())
        
    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('reportPage'))

# Export Report Routes
@app.route('/export-report/<format>')
@login_required
@role_required('HR')
def export_report(format):
    """Export current report data in specified format"""
    
    # Get the report type from query parameter or default to employee roster
    report_type = request.args.get('type', 'employee-roster')
    
    if format == 'csv':
        return export_to_csv(report_type)
    elif format == 'excel':
        if not OPENPYXL_AVAILABLE:
            flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'info')
            return export_to_csv(report_type)
        return export_to_excel(report_type)
    elif format == 'pdf':
        if not REPORTLAB_AVAILABLE:
            flash('PDF export requires reportlab. Install with: pip install reportlab', 'info')
            flash('Please use Print option or download CSV instead.', 'info')
            return redirect(url_for('reportPage'))
        return export_to_pdf(report_type)
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reportPage'))

@app.route('/export-report/<format>/<report_type>')
@login_required
@role_required('HR')
def export_report_new(format, report_type):
    """Export reports in PDF, Excel, or CSV format"""
    
    if format == 'pdf':
        if not REPORTLAB_AVAILABLE:
            flash('PDF export requires reportlab. Install with: pip install reportlab', 'error')
            return redirect(url_for('reportPage'))
        return export_to_pdf(report_type)
    
    elif format == 'excel':
        if not OPENPYXL_AVAILABLE:
            flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'error')
            return redirect(url_for('reportPage'))
        return export_to_excel(report_type)
    
    elif format == 'csv':
        return export_to_csv(report_type)
    
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reportPage'))


# ============================================
# PDF EXPORT FUNCTION
# ============================================

def export_to_pdf(report_type):
    """Generate PDF report"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#2d7fa6'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add SwiftSend logo/header
        elements.append(Paragraph("SwiftSend Couriers", title_style))
        elements.append(Paragraph("Human Resources Management System", subtitle_style))
        
        if report_type == 'employee-roster':
            cursor.execute("""
                SELECT employee_id, first_name, last_name, position, 
                       email, phone, hire_date
                FROM Employees
                ORDER BY last_name, first_name
            """)
            employees = cursor.fetchall()
            
            elements.append(Paragraph("Employee Roster Report", styles['Heading2']))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Table data
            data = [['ID', 'Name', 'Position', 'Email', 'Phone', 'Hire Date']]
            for emp in employees:
                data.append([
                    str(emp['employee_id']),
                    f"{emp['first_name']} {emp['last_name']}",
                    emp['position'][:25],  # Truncate long positions
                    emp['email'][:30],
                    emp['phone'] or 'N/A',
                    emp['hire_date'].strftime('%Y-%m-%d') if emp['hire_date'] else 'N/A'
                ])
            
            # Create table with adjusted widths
            table = Table(data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4da6d1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
            
            # Add footer
            elements.append(Spacer(1, 0.5*inch))
            elements.append(Paragraph(f"Total Employees: {len(employees)}", styles['Normal']))
            
        elif report_type == 'leave-summary':
            cursor.execute("""
                SELECT e.first_name, e.last_name, e.position, lr.leave_type, 
                       lr.start_date, lr.end_date, lr.days_requested, lr.status, lr.reason
                FROM LeaveRequests lr
                JOIN Employees e ON lr.employee_id = e.employee_id
                ORDER BY lr.start_date DESC
                LIMIT 100
            """)
            leaves = cursor.fetchall()
            
            elements.append(Paragraph("Leave Summary Report", styles['Heading2']))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            data = [['Employee', 'Position', 'Type', 'Start', 'End', 'Days', 'Status']]
            for leave in leaves:
                data.append([
                    f"{leave['first_name']} {leave['last_name']}"[:20],
                    leave['position'][:15],
                    leave['leave_type'],
                    leave['start_date'].strftime('%Y-%m-%d'),
                    leave['end_date'].strftime('%Y-%m-%d'),
                    str(int(leave['days_requested'])),
                    leave['status']
                ])
            
            table = Table(data, colWidths=[1.3*inch, 1.2*inch, 0.8*inch, 0.9*inch, 0.9*inch, 0.6*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4da6d1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
        
        elif report_type == 'document-compliance':
            cursor.execute("""
                SELECT e.employee_id, e.first_name, e.last_name, e.position,
                       COUNT(d.document_id) as total_docs,
                       SUM(CASE WHEN d.expiry_date < CURDATE() THEN 1 ELSE 0 END) as expired,
                       SUM(CASE WHEN d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) 
                           THEN 1 ELSE 0 END) as expiring
                FROM Employees e
                LEFT JOIN Documents d ON e.employee_id = d.employee_id
                GROUP BY e.employee_id
                ORDER BY e.last_name, e.first_name
            """)
            compliance = cursor.fetchall()
            
            elements.append(Paragraph("Document Compliance Report", styles['Heading2']))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            data = [['ID', 'Employee', 'Position', 'Total Docs', 'Expired', 'Expiring', 'Status']]
            for comp in compliance:
                expired = comp['expired'] or 0
                expiring = comp['expiring'] or 0
                status = 'Non-Compliant' if expired > 0 else ('At Risk' if expiring > 0 else 'Compliant')
                
                data.append([
                    str(comp['employee_id']),
                    f"{comp['first_name']} {comp['last_name']}"[:20],
                    comp['position'][:18],
                    str(comp['total_docs']),
                    str(expired),
                    str(expiring),
                    status
                ])
            
            table = Table(data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 0.9*inch, 0.8*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4da6d1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
        
        cursor.close()
        conn.close()
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'swiftsend_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('reportPage'))


# ============================================
# EXCEL EXPORT FUNCTION
# ============================================

def export_to_excel(report_type):
    """Generate Excel report"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        
        # Styles
        header_fill = PatternFill(start_color='4da6d1', end_color='4da6d1', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if report_type == 'employee-roster':
            ws.title = "Employee Roster"
            
            cursor.execute("""
                SELECT employee_id, first_name, last_name, position, 
                       email, phone, hire_date, annual_leave_days, sick_leave_days
                FROM Employees
                ORDER BY last_name, first_name
            """)
            employees = cursor.fetchall()
            
            # Headers
            headers = ['ID', 'First Name', 'Last Name', 'Position', 'Email', 'Phone', 'Hire Date', 'Annual Leave', 'Sick Leave']
            ws.append(headers)
            
            # Style headers
            for col_num, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            for emp in employees:
                ws.append([
                    emp['employee_id'],
                    emp['first_name'],
                    emp['last_name'],
                    emp['position'],
                    emp['email'],
                    emp['phone'] or 'N/A',
                    emp['hire_date'].strftime('%Y-%m-%d') if emp['hire_date'] else 'N/A',
                    emp['annual_leave_days'],
                    emp['sick_leave_days']
                ])
            
        elif report_type == 'leave-summary':
            ws.title = "Leave Summary"
            
            cursor.execute("""
                SELECT e.first_name, e.last_name, e.position, lr.leave_type, 
                       lr.start_date, lr.end_date, lr.days_requested, lr.status, lr.reason
                FROM LeaveRequests lr
                JOIN Employees e ON lr.employee_id = e.employee_id
                ORDER BY lr.start_date DESC
            """)
            leaves = cursor.fetchall()
            
            headers = ['First Name', 'Last Name', 'Position', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Status', 'Reason']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for leave in leaves:
                ws.append([
                    leave['first_name'],
                    leave['last_name'],
                    leave['leave_type'],
                    leave['start_date'].strftime('%Y-%m-%d'),
                    leave['end_date'].strftime('%Y-%m-%d'),
                    float(leave['days_requested']),
                    leave['status'],
                    leave['reason'] or 'N/A'
                ])
        
        elif report_type == 'document-compliance':
            ws.title = "Document Compliance"
            
            cursor.execute("""
                SELECT e.employee_id, e.first_name, e.last_name, e.position,
                       COUNT(d.document_id) as total_docs,
                       SUM(CASE WHEN d.expiry_date < CURDATE() THEN 1 ELSE 0 END) as expired,
                       SUM(CASE WHEN d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) 
                           THEN 1 ELSE 0 END) as expiring_soon
                FROM Employees e
                LEFT JOIN Documents d ON e.employee_id = d.employee_id
                GROUP BY e.employee_id
                ORDER BY e.last_name, e.first_name
            """)
            compliance = cursor.fetchall()
            
            headers = ['Employee ID', 'First Name', 'Last Name', 'Position', 'Total Docs', 'Expired', 'Expiring Soon', 'Compliance Status']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for comp in compliance:
                expired = comp['expired'] or 0
                expiring = comp['expiring_soon'] or 0
                status = 'Non-Compliant' if expired > 0 else ('At Risk' if expiring > 0 else 'Compliant')
                
                ws.append([
                    comp['employee_id'],
                    comp['first_name'],
                    comp['last_name'],
                    comp['position'],
                    comp['total_docs'],
                    expired,
                    expiring,
                    status
                ])
        
        cursor.close()
        conn.close()
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    cell.border = border
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'swiftsend_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('reportPage'))


# ============================================
# CSV EXPORT FUNCTION
# ============================================

def export_to_csv(report_type):
    """Generate CSV report"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Create CSV in memory
        output = StringIO()
        
        if report_type == 'employee-roster':
            cursor.execute("""
                SELECT employee_id, first_name, last_name, position, 
                       email, phone, hire_date
                FROM Employees
                ORDER BY last_name, first_name
            """)
            employees = cursor.fetchall()
            
            writer = csv.writer(output)
            writer.writerow(['Employee ID', 'First Name', 'Last Name', 'Position', 'Email', 'Phone', 'Hire Date'])
            
            for emp in employees:
                writer.writerow([
                    emp['employee_id'],
                    emp['first_name'],
                    emp['last_name'],
                    emp['position'],
                    emp['email'],
                    emp['phone'] or 'N/A',
                    emp['hire_date'].strftime('%Y-%m-%d') if emp['hire_date'] else 'N/A'
                ])
        
        elif report_type == 'leave-summary':
            cursor.execute("""
                SELECT e.first_name, e.last_name, lr.leave_type, 
                       lr.start_date, lr.end_date, lr.days_requested, lr.status
                FROM LeaveRequests lr
                JOIN Employees e ON lr.employee_id = e.employee_id
                ORDER BY lr.start_date DESC
            """)
            leaves = cursor.fetchall()
            
            writer = csv.writer(output)
            writer.writerow(['First Name', 'Last Name', 'Leave Type', 'Start Date', 'End Date', 'Days Requested', 'Status'])
            
            for leave in leaves:
                writer.writerow([
                    leave['first_name'],
                    leave['last_name'],
                    leave['leave_type'],
                    leave['start_date'].strftime('%Y-%m-%d'),
                    leave['end_date'].strftime('%Y-%m-%d'),
                    leave['days_requested'],
                    leave['status']
                ])
        
        elif report_type == 'document-compliance':
            cursor.execute("""
                SELECT e.employee_id, e.first_name, e.last_name, e.position,
                       COUNT(d.document_id) as total_docs,
                       SUM(CASE WHEN d.expiry_date < CURDATE() THEN 1 ELSE 0 END) as expired,
                       SUM(CASE WHEN d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) 
                           THEN 1 ELSE 0 END) as expiring
                FROM Employees e
                LEFT JOIN Documents d ON e.employee_id = d.employee_id
                GROUP BY e.employee_id
                ORDER BY e.last_name, e.first_name
            """)
            compliance = cursor.fetchall()
            
            writer = csv.writer(output)
            writer.writerow(['Employee ID', 'First Name', 'Last Name', 'Position', 'Total Docs', 'Expired', 'Expiring Soon', 'Status'])
            
            for comp in compliance:
                expired = comp['expired'] or 0
                expiring = comp['expiring'] or 0
                status = 'Non-Compliant' if expired > 0 else ('At Risk' if expiring > 0 else 'Compliant')
                
                writer.writerow([
                    comp['employee_id'],
                    comp['first_name'],
                    comp['last_name'],
                    comp['position'],
                    comp['total_docs'],
                    expired,
                    expiring,
                    status
                ])
        
        cursor.close()
        conn.close()
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=swiftsend_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        response.headers["Content-type"] = "text/csv"
        
        return response
        
    except Exception as e:
        flash(f'Error generating CSV: {str(e)}', 'error')
        return redirect(url_for('reportPage'))

def export_csv(data):
    """Export data to CSV format"""
    si = StringIO()
    writer = csv.writer(si)
    
    # Write header
    if data:
        writer.writerow(data[0].keys())
        
        # Write data rows
        for row in data:
            writer.writerow(row.values())
    
    # Create response
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=hr_report_{datetime.now().strftime('%Y%m%d')}.csv"
    output.headers["Content-type"] = "text/csv"
    
    return output

@app.route('/performance')
@login_required
@role_required('HR')
def performance():
    """HR - View employee performance and compliance metrics"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get comprehensive employee metrics
        cursor.execute("""
            SELECT 
                e.employee_id,
                e.first_name,
                e.last_name,
                e.position,
                e.hire_date,
                e.annual_leave_days,
                e.leave_days_used,
                TIMESTAMPDIFF(YEAR, e.hire_date, CURDATE()) as years_of_service,
                COUNT(DISTINCT d.document_id) as total_docs,
                SUM(CASE WHEN d.expiry_date < CURDATE() THEN 1 ELSE 0 END) as expired_docs,
                SUM(CASE WHEN d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 1 ELSE 0 END) as expiring_docs
            FROM Employees e
            LEFT JOIN Documents d ON e.employee_id = d.employee_id
            GROUP BY e.employee_id
            ORDER BY e.last_name, e.first_name
        """)
        employees = cursor.fetchall()
        
        # Calculate summary statistics
        total_employees = len(employees)
        compliant_employees = sum(1 for emp in employees if emp['expired_docs'] == 0 and emp['expiring_docs'] == 0)
        non_compliant = sum(1 for emp in employees if emp['expired_docs'] > 0 or emp['expiring_docs'] > 0)
        
        # Calculate average leave utilization
        if total_employees > 0:
            total_utilization = sum((emp['leave_days_used'] / emp['annual_leave_days'] * 100) 
                                   for emp in employees if emp['annual_leave_days'] > 0)
            avg_leave_utilization = round(total_utilization / total_employees, 1)
        else:
            avg_leave_utilization = 0
        
        # Get department distribution
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN position LIKE '%Driver%' OR position LIKE '%Courier%' THEN 'Drivers'
                    WHEN position LIKE '%Manager%' OR position LIKE '%Supervisor%' THEN 'Management'
                    WHEN position LIKE '%HR%' THEN 'Human Resources'
                    WHEN position LIKE '%Admin%' OR position LIKE '%Assistant%' THEN 'Administration'
                    WHEN position LIKE '%Customer%' THEN 'Customer Service'
                    ELSE 'Other'
                END as department,
                COUNT(*) as count
            FROM Employees
            GROUP BY department
            ORDER BY count DESC
        """)
        departments = cursor.fetchall()
        
        # Calculate percentages
        for dept in departments:
            dept['percentage'] = round((dept['count'] / total_employees * 100), 1) if total_employees > 0 else 0
        
        # Get longest tenure employee
        cursor.execute("""
            SELECT first_name, last_name, 
                   TIMESTAMPDIFF(YEAR, hire_date, CURDATE()) as years
            FROM Employees
            ORDER BY hire_date ASC
            LIMIT 1
        """)
        longest_tenure = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return render_template(
            'performance.html',
            user=session,
            employees=employees,
            total_employees=total_employees,
            compliant_employees=compliant_employees,
            non_compliant=non_compliant,
            avg_leave_utilization=avg_leave_utilization,
            departments=departments,
            longest_tenure=longest_tenure
        )
        
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect(url_for('hr_dashboard'))

# ENHANCED MONTHLY SUMMARY REPORT
@app.route('/generate-report/monthly_summary') # Changed to use the route parameter directly for clarity
@login_required
@role_required('HR')
def generate_monthly_summary_report():
    """Generate the enhanced monthly summary report"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get current month and year
        current_month = datetime.now().strftime('%Y-%m')
        current_date = datetime.now()
        
        # 1. EMPLOYEE SUMMARY
        # Total employees
        cursor.execute("SELECT COUNT(*) as total_employees FROM Employees")
        total_employees = cursor.fetchone()['total_employees']
        
        # New hires this month
        cursor.execute("""
            SELECT COUNT(*) as new_hires,
                   GROUP_CONCAT(CONCAT(first_name, ' ', last_name) SEPARATOR ', ') as new_hire_names
            FROM Employees
            WHERE DATE_FORMAT(hire_date, '%%Y-%%m') = %s
        """, (current_month,))
        new_hires_data = cursor.fetchone()
        
        # Employees with contracts expiring this month or already expired
        cursor.execute("""
            SELECT COUNT(DISTINCT e.employee_id) as contract_updates,
                   GROUP_CONCAT(DISTINCT CONCAT(e.first_name, ' ', e.last_name, ' (', 
                                DATE_FORMAT(d.expiry_date, '%%d %%b %%Y'), ')') SEPARATOR ', ') as employee_names
            FROM Employees e
            JOIN Documents d ON e.employee_id = d.employee_id
            WHERE d.doc_type = 'Contract' 
            AND (d.expiry_date < CURDATE() OR DATE_FORMAT(d.expiry_date, '%%Y-%%m') = %s)
        """, (current_month,))
        contract_updates = cursor.fetchone()
        
        # 2. LEAVE SUMMARY
        # Total leave requests this month
        cursor.execute("""
            SELECT COUNT(*) as total_leaves,
                   SUM(CASE WHEN status = 'Approved' THEN 1 ELSE 0 END) as approved,
                   SUM(CASE WHEN status = 'Rejected' THEN 1 ELSE 0 END) as rejected,
                   SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending
            FROM LeaveRequests
            WHERE DATE_FORMAT(start_date, '%%Y-%%m') = %s
        """, (current_month,))
        leave_stats = cursor.fetchone()
        
        # Most common leave types
        cursor.execute("""
            SELECT leave_type, COUNT(*) as count
            FROM LeaveRequests
            WHERE DATE_FORMAT(start_date, '%%Y-%%m') = %s
            GROUP BY leave_type
            ORDER BY count DESC
        """, (current_month,))
        leave_types = cursor.fetchall()
        
        # Employees currently on leave
        cursor.execute("""
            SELECT CONCAT(e.first_name, ' ', e.last_name) as employee_name,
                   lr.leave_type,
                   lr.start_date,
                   lr.end_date
            FROM LeaveRequests lr
            JOIN Employees e ON lr.employee_id = e.employee_id
            WHERE lr.status = 'Approved'
            AND CURDATE() BETWEEN lr.start_date AND lr.end_date
            ORDER BY lr.end_date
        """)
        current_leaves = cursor.fetchall()
        
        # 3. DOCUMENT EXPIRY ALERTS
        # Documents expiring within 60 days
        cursor.execute("""
            SELECT CONCAT(e.first_name, ' ', e.last_name) as employee_name,
                   d.doc_type,
                   d.expiry_date,
                   DATEDIFF(d.expiry_date, CURDATE()) as days_until_expiry
            FROM Documents d
            JOIN Employees e ON d.employee_id = e.employee_id
            WHERE d.expiry_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 60 DAY)
            ORDER BY d.expiry_date
        """)
        expiring_docs = cursor.fetchall()
        
        # 4. PERFORMANCE INDICATORS
        # Average leave days used per employee
        cursor.execute("""
            SELECT AVG(leave_days_used) as avg_leave_days
            FROM Employees
        """)
        avg_leave = cursor.fetchone()
        
        # Department/Position with most leave activity
        cursor.execute("""
            SELECT e.position, COUNT(lr.leave_id) as leave_count
            FROM LeaveRequests lr
            JOIN Employees e ON lr.employee_id = e.employee_id
            WHERE DATE_FORMAT(lr.start_date, '%%Y-%%m') = %s
            GROUP BY e.position
            ORDER BY leave_count DESC
            LIMIT 1
        """, (current_month,))
        top_leave_dept = cursor.fetchone()
        
        # Count of unread expiry alerts
        cursor.execute("""
            SELECT COUNT(*) as unread_alerts
            FROM ExpiryAlerts
            WHERE status = 'Unread'
        """)
        unread_alerts = cursor.fetchone()
        
        # Compile all data
        data = {
            'total_employees': total_employees,
            'new_hires': new_hires_data['new_hires'] or 0,
            'new_hire_names': new_hires_data['new_hire_names'] or 'None',
            'contract_updates': contract_updates['contract_updates'] or 0,
            'contract_update_names': contract_updates['employee_names'] or 'None',
            'leave_stats': leave_stats,
            'leave_types': leave_types,
            'current_leaves': current_leaves,
            'expiring_docs': expiring_docs,
            'avg_leave_days': round(avg_leave['avg_leave_days'] or 0, 2),
            'top_leave_dept': top_leave_dept,
            'unread_alerts': unread_alerts['unread_alerts'] or 0,
            'month': datetime.now().strftime('%B %Y')
        }
        title = f"Monthly Summary Report - {datetime.now().strftime('%B %Y')}"
        template = 'reports/monthly_summary.html'
        
        cursor.close()
        conn.close()

        return render_template(template, data=data, title=title, generated_date=datetime.now())

    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('reportPage'))
    except Exception as e:
        flash(f'An unexpected error occurred: {str(e)}', 'error')
        return redirect(url_for('reportPage'))

if __name__ == "__main__":
    app.run(debug=True)
